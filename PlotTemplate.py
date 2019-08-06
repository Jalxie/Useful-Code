#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Sep  6 11:32:05 2018

@author: jalxiey
"""
import numpy as np 
import matplotlib.pyplot as plt # For Plotting
import xlrd # For import data from .xlsx file
import pandas as pd
from openpyxl import load_workbook
from pylab import figure, show, legend, ylabel

#data to use
data1 = []
data2 = []
data3 = []


xl1 = pd.ExcelFile('R0.39_4_2V_NoRing.xlsx', sep='\t')
df1 = xl1.parse('Sheet1')
df2 = xl1.parse('Sheet2')

# Load in the workbook
wb1 = load_workbook('R0.39_4_2V_NoRing.xlsx')

# Get sheet names
print(wb1.get_sheet_names())
print('\n')
# Get a sheet by name 
sheet1 = wb1.get_sheet_by_name('Sheet1')

print(sheet1['A1'].value)

# Get a sheet by name 
sheet2 = wb1.get_sheet_by_name('Sheet2')

print(sheet2['A1'].value)



#print all the column name
for i in range(1, sheet1.max_column):
     print(i, sheet1.cell(row=1, column=i).value)

for i in range(3, sheet1.max_row):
    if isinstance(sheet1.cell(row=i, column=4).value,float):
        data1.append(sheet1.cell(row=i, column=4).value)
        
for i in range(3, sheet1.max_row):
    if isinstance(sheet1.cell(row=i, column=6).value,float):
        data2.append(sheet1.cell(row=i, column=4).value)
        
        #print all the column name
for i in range(1, sheet2.max_column):
     print(i, sheet2.cell(row=1, column=i).value)

for i in range(1, sheet2.max_row):
    if isinstance(sheet2.cell(row=i, column=8).value,float):
        data3.append(sheet2.cell(row=i, column=8).value)
        


"""
Plotting
"""
plt.figure(1).add_subplot(111)

plt.plot(data1)
plt.plot(data2)
plt.plot(data3)


# create the general figure
fig1 = figure()
 
# and the first axes using subplot populated with data 
ax1 = fig1.add_subplot(111)
line1 = ax1.plot(data1)
line2 = ax1.plot(data2)
ylabel("Temperature")
 
# now, the second axes that shares the x-axis with the ax1
ax2 = fig1.add_subplot(111, sharex=ax1, frameon=False)
line3 = ax2.plot(data3)
ax2.yaxis.tick_right()
ax2.yaxis.set_label_position("right")
ylabel("Lumens")
 
# for the legend, remember that we used two different axes so, we need 
# to build the legend manually
plt.legend((line1, line2, line3), ("T1", "T2", "Lumens"))
show()

"""
#import data from excel
ExcelFileName= 'R0.39_4_2V_NoRing.xlsx'
workbook = xlrd.open_workbook(ExcelFileName)
worksheet = workbook.sheet_by_name("Sheet1") # We need to read the data 
#from the Excel sheet named "Sheet1"
num_rows = worksheet.nrows #Number of Rows
num_cols = worksheet.ncols #Number of Columns

result_data =[]
for curr_row in range(0, num_rows, 1):
 row_data = []
 for curr_col in range(0, num_cols, 1):
     data = worksheet.cell_value(curr_row, curr_col) # Read the data in the current cell
     print(data)
     row_data.append(data)
     result_data.append(row_data)
"""
"""
#plotting
plt.plot([1,2,3,4], [1,2,3,4])

#label the axises
plt.xlabel('some numbers')
plt.ylabel('some numbers')

#define the range of axis
plt.axis([0, 6, 0, 20])

#show the plot in the console
plt.show()
"""