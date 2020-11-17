import numpy as np 
import matplotlib.pyplot as plt # For Plotting
import xlrd # For import data from .xlsx file
import pandas as pd
from openpyxl import load_workbook
from pylab import figure, show, legend, ylabel

ydata = []
xdata = []
filename = 'filename.xlsx'
sheetname = 'Sheet1'

# Load in the workbook
wb = load_workbook(filename)

# Get sheet names
print(wb.get_sheet_names())
print('\n')
# Get a sheet by name 
sheet = wb.get_sheet_by_name(sheetname)

#print all the column name
# for i in range(1, 5120):
#      print(i, sheet.cell(row=i, column=1).value)

for i in range(2, 5120):
    xdata.append(sheet.cell(row=i, column=1).value)

# for i in range(1, 5120):
#      print(i, sheet.cell(row=i, column=2).value)

for i in range(2, 5120): 
    # if sheet.cell(row=i, column=2).value > 0:
    #     ydata.append(1000)
    # else: 
    #     ydata.append(0)
    ydata.append(sheet.cell(row=i, column=2).value)
"""
Plotting
"""
plt.figure(1)

plt.plot(xdata[0:1024*6:1], ydata[0:1024*6:1])

plt.title("40Hz")

plt.show()
